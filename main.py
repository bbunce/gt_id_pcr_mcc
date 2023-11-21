# import openpyxl library
import openpyxl

# create a workbook object
wb = openpyxl.load_workbook('data/example_workbook.xlsx')
# create a worksheet object
ws = wb['Sheet1']

# print the value of a specific cell
print(ws.cell(row=1, column=1).value)

# for loop to iterate over rows
for row in range(2, 6):
    cell = ws.cell(row=row, column=1).value
    print(cell)
